import json
import datetime
import random
from django.utils import timezone
from datetime import timedelta
from django.core.mail import send_mail, EmailMultiAlternatives, get_connection
from email.mime.image import MIMEImage
import os
import threading
from django.conf import settings
from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
import openpyxl
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from django.contrib import messages
from django.db import transaction
from .models import Docente, Curso, Paralelo, Horario, Usuario, Asignatura, Leccion, DisponibilidadDocente, SesionGenerador, BorradorHorario, Aula
from .forms import HorarioForm, HorarioAdminForm, DocenteEditForm, DocenteCreateForm, CursoForm, ParaleloForm, AsignaturaForm, DocenteFotoForm, AtencionPadresForm
from .forms_scheduler import LeccionForm


def admin_required(view_func):
    """Decorator: solo administradores. Redirige a index con mensaje si no es admin."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('horarios:login')
        if not (request.user.is_admin() or request.user.is_superuser):
            messages.warning(request, 'Solo se permite acceso a administradores')
            return redirect('horarios:index')
        return view_func(request, *args, **kwargs)
    return wrapper

TIME_TO_ROW = {
    "07:00:00": 2,
    "07:45:00": 3,
    "08:30:00": 4,
    "09:15:00": 5,
    "10:00:00": 6,
    "10:25:00": 7,
    "11:05:00": 8,
    "11:45:00": 9,
    "12:25:00": 10,
    "13:05:00": 11,
    "13:45:00": 12,
}

DAY_TO_COL = {
    "lunes": 2, "martes": 3, "miercoles": 4, "jueves": 5, "viernes": 6,
}

PERIODOS_ROWS = [
    (2, "07:00:00", "07:45:00"),
    (3, "07:45:00", "08:30:00"),
    (4, "08:30:00", "09:15:00"),
    (5, "09:15:00", "10:00:00"),
    (7, "10:25:00", "11:05:00"),
    (8, "11:05:00", "11:45:00"),
    (9, "11:45:00", "12:25:00"),
    (10, "12:25:00", "13:05:00"),
    (11, "13:05:00", "13:45:00"),
]


# Paleta ampliada: 36 colores con hues bien distribuidos (cada 10°)
# para que sea casi imposible que dos combinaciones distintas repitan color.
_PALETA = [
    {'bg': 'hsl(0, 80%, 45%)', 'border': 'hsl(0, 90%, 30%)', 'text': '#ffffff'},
    {'bg': 'hsl(210, 85%, 70%)', 'border': 'hsl(210, 80%, 40%)', 'text': 'hsl(210, 90%, 15%)'},
    {'bg': 'hsl(120, 80%, 35%)', 'border': 'hsl(120, 90%, 20%)', 'text': '#ffffff'},
    {'bg': 'hsl(45, 90%, 65%)', 'border': 'hsl(45, 85%, 35%)', 'text': 'hsl(45, 90%, 15%)'},
    {'bg': 'hsl(280, 80%, 50%)', 'border': 'hsl(280, 90%, 30%)', 'text': '#ffffff'},
    {'bg': 'hsl(15, 85%, 65%)', 'border': 'hsl(15, 80%, 35%)', 'text': 'hsl(15, 90%, 15%)'},
    {'bg': 'hsl(180, 85%, 40%)', 'border': 'hsl(180, 90%, 25%)', 'text': '#ffffff'},
    {'bg': 'hsl(330, 85%, 70%)', 'border': 'hsl(330, 80%, 40%)', 'text': 'hsl(330, 90%, 15%)'},
    {'bg': 'hsl(80, 80%, 45%)', 'border': 'hsl(80, 90%, 25%)', 'text': '#ffffff'},
    {'bg': 'hsl(240, 80%, 65%)', 'border': 'hsl(240, 80%, 40%)', 'text': 'hsl(240, 90%, 15%)'},
    {'bg': 'hsl(30, 90%, 50%)', 'border': 'hsl(30, 90%, 30%)', 'text': '#ffffff'},
    {'bg': 'hsl(300, 85%, 70%)', 'border': 'hsl(300, 80%, 40%)', 'text': 'hsl(300, 90%, 15%)'},
    {'bg': 'hsl(150, 80%, 35%)', 'border': 'hsl(150, 90%, 20%)', 'text': '#ffffff'},
    {'bg': 'hsl(260, 85%, 65%)', 'border': 'hsl(260, 80%, 40%)', 'text': 'hsl(260, 90%, 15%)'},
    {'bg': 'hsl(345, 80%, 50%)', 'border': 'hsl(345, 90%, 30%)', 'text': '#ffffff'},
    {'bg': 'hsl(60, 85%, 65%)', 'border': 'hsl(60, 80%, 35%)', 'text': 'hsl(60, 90%, 15%)'},
    {'bg': 'hsl(195, 80%, 45%)', 'border': 'hsl(195, 90%, 25%)', 'text': '#ffffff'},
    {'bg': 'hsl(315, 85%, 70%)', 'border': 'hsl(315, 80%, 40%)', 'text': 'hsl(315, 90%, 15%)'},
    {'bg': 'hsl(10, 80%, 45%)', 'border': 'hsl(10, 90%, 25%)', 'text': '#ffffff'},
    {'bg': 'hsl(220, 85%, 65%)', 'border': 'hsl(220, 80%, 40%)', 'text': 'hsl(220, 90%, 15%)'},
    {'bg': 'hsl(100, 80%, 35%)', 'border': 'hsl(100, 90%, 20%)', 'text': '#ffffff'},
    {'bg': 'hsl(40, 90%, 65%)', 'border': 'hsl(40, 85%, 35%)', 'text': 'hsl(40, 90%, 15%)'},
    {'bg': 'hsl(270, 80%, 50%)', 'border': 'hsl(270, 90%, 30%)', 'text': '#ffffff'},
    {'bg': 'hsl(350, 85%, 70%)', 'border': 'hsl(350, 80%, 40%)', 'text': 'hsl(350, 90%, 15%)'},
    {'bg': 'hsl(170, 80%, 35%)', 'border': 'hsl(170, 90%, 20%)', 'text': '#ffffff'},
    {'bg': 'hsl(230, 85%, 65%)', 'border': 'hsl(230, 80%, 40%)', 'text': 'hsl(230, 90%, 15%)'},
    {'bg': 'hsl(70, 80%, 40%)', 'border': 'hsl(70, 90%, 25%)', 'text': '#ffffff'},
    {'bg': 'hsl(320, 85%, 70%)', 'border': 'hsl(320, 80%, 40%)', 'text': 'hsl(320, 90%, 15%)'},
    {'bg': 'hsl(140, 80%, 35%)', 'border': 'hsl(140, 90%, 20%)', 'text': '#ffffff'},
    {'bg': 'hsl(250, 85%, 65%)', 'border': 'hsl(250, 80%, 40%)', 'text': 'hsl(250, 90%, 15%)'},
    {'bg': 'hsl(20, 80%, 45%)', 'border': 'hsl(20, 90%, 25%)', 'text': '#ffffff'},
    {'bg': 'hsl(200, 85%, 65%)', 'border': 'hsl(200, 80%, 40%)', 'text': 'hsl(200, 90%, 15%)'}
]

def get_color_for_materia(clave):
    """
    Devuelve un color de la paleta según la clave dada.
    La clave debe ser materia|curso_id|paralelo_id para que
    el mismo grupo siempre tenga el mismo color.
    Usa avalanche mixing para dispersar mejor strings similares.
    """
    hash_val = 5381
    for c in str(clave).lower().strip():
        # djb2 con avalanche: mejor dispersión para strings parecidos
        hash_val = ((hash_val << 5) + hash_val + ord(c)) & 0xFFFFFFFF
    # Segunda pasada para mejorar la dispersión de bits
    hash_val ^= (hash_val >> 16)
    hash_val = (hash_val * 0x45d9f3b) & 0xFFFFFFFF
    hash_val ^= (hash_val >> 16)
    return _PALETA[hash_val % len(_PALETA)]



def preparar_horarios_grid(horarios_qs, incluir_vacios=False, fusionar_bloques=False, color_by='default'):
    base_items = []
    filled_map = {}
    
    for h in horarios_qs:
        start_str = h.hora_inicio.strftime("%H:%M:%S")
        end_str = h.hora_fin.strftime("%H:%M:%S")
        
        row_start = TIME_TO_ROW.get(start_str, 2)
        row_end = TIME_TO_ROW.get(end_str, row_start + 1)
        col = DAY_TO_COL.get(h.dia.lower(), 2)

        if color_by == 'curso':
            color_index = (h.paralelo_id * 7) if h.paralelo_id else 0
        elif color_by == 'asignatura':
            color_index = (h.asignatura_id * 7) if h.asignatura_id else 0
        else:
            color_index = (h.asignatura_id or 0) * 11 + (h.paralelo_id or 0) * 7
            
        colors = _PALETA[color_index % len(_PALETA)]
        
        base_items.append({
            'horario': h,
            'grid_row_start': row_start,
            'grid_row_end': row_end,
            'grid_col': col,
            'color_bg':     colors['bg'],
            'color_border': colors['border'],
            'color_text':   colors['text'],
        })
        
        for r in range(row_start, row_end):
            filled_map[(col, r)] = True

    if fusionar_bloques:
        base_items.sort(key=lambda x: (x['grid_col'], x['grid_row_start']))
        merged_items = []
        for item in base_items:
            if not merged_items:
                merged_items.append(item)
                continue
            last = merged_items[-1]
            can_merge = (
                last['grid_col'] == item['grid_col'] and
                last['grid_row_end'] == item['grid_row_start'] and
                last['horario'].asignatura_id == item['horario'].asignatura_id and
                last['horario'].docente_id == item['horario'].docente_id and
                last['horario'].curso_id == item['horario'].curso_id and
                last['horario'].paralelo_id == item['horario'].paralelo_id
            )
            if can_merge:
                last['grid_row_end'] = item['grid_row_end']
            else:
                merged_items.append(item)
        horarios_list = merged_items
    else:
        horarios_list = base_items

    if incluir_vacios:
        celdas_vacias = []
        for dia, col in DAY_TO_COL.items():
            for row, h_inicio, h_fin in PERIODOS_ROWS:
                celdas_vacias.append({
                    'dia': dia,
                    'hora_inicio': h_inicio,
                    'hora_fin': h_fin,
                    'grid_col': col,
                    'grid_row': row
                })
        return horarios_list, celdas_vacias
        
    return horarios_list

def login_view(request):
    if request.user.is_authenticated:
        next_url = request.GET.get('next')
        if next_url:
            return redirect(next_url)
        if request.user.is_docente():
            return redirect('horarios:dashboard_docente')
        elif request.user.is_admin() or request.user.is_superuser:
            return redirect('horarios:gestion_dashboard')
        else:
            return redirect('horarios:index')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if not user.is_superuser and not user.email.endswith('@ladolorosa-loja.edu.ec'):
                messages.warning(request, 'El correo asociado a esta cuenta no pertenece a la institución (@ladolorosa-loja.edu.ec). Acceso denegado.')
                return render(request, 'horarios/login.html')
                
            if user.two_factor_enabled:
                code = str(random.randint(100000, 999999))
                user.two_factor_code = code
                user.two_factor_expires = timezone.now() + timedelta(minutes=10)
                user.save()
                
                html_content = render_to_string('horarios/emails/2fa_code.html', {
                    'username': user.first_name or user.username,
                    'code': code
                })
                text_content = f"Tu código de inicio de sesión es: {code}. Expirará en 10 minutos."
                
                from_email_formatted = f"GeoCampus La Dolorosa <{settings.EMAIL_HOST_USER}>"
                msg = EmailMultiAlternatives(
                    subject="Tu código de verificación - La Dolorosa",
                    body=text_content,
                    from_email=from_email_formatted,
                    to=[user.email],
                    reply_to=[settings.EMAIL_HOST_USER]
                )
                msg.attach_alternative(html_content, "text/html")
                
                # Adjuntar imagen
                logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_dolorosa.png')
                if os.path.exists(logo_path):
                    with open(logo_path, 'rb') as f:
                        logo_img = MIMEImage(f.read())
                        logo_img.add_header('Content-ID', '<logo>')
                        logo_img.add_header('Content-Disposition', 'inline', filename='logo_dolorosa.png')
                        msg.attach(logo_img)
                        
                msg.send(fail_silently=False)
                
                request.session['pre_2fa_user_id'] = user.id
                return redirect('horarios:verify_2fa')
                
            login(request, user)
            
            next_url = request.POST.get('next') or request.GET.get('next')
            if next_url:
                return redirect(next_url)
                
            if user.is_docente():
                return redirect('horarios:dashboard_docente')
            elif user.is_admin() or user.is_superuser:
                return redirect('horarios:gestion_dashboard')
            else:
                return redirect('horarios:index')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'horarios/login.html')

def _send_registro_code(email, first_name, code):
    """Envía el correo de verificación de registro."""
    html_content = render_to_string('horarios/emails/registro_verificacion.html', {
        'username': first_name,
        'code': code,
    })
    text_content = f"Tu código de verificación para crear tu cuenta es: {code}. Expirará en 15 minutos."
    from_email_formatted = f"GeoCampus La Dolorosa <{settings.EMAIL_HOST_USER}>"
    msg = EmailMultiAlternatives(
        subject="Código de verificación para tu nueva cuenta - La Dolorosa",
        body=text_content,
        from_email=from_email_formatted,
        to=[email],
        reply_to=[settings.EMAIL_HOST_USER],
    )
    msg.attach_alternative(html_content, "text/html")
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_dolorosa.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_img = MIMEImage(f.read())
            logo_img.add_header('Content-ID', '<logo>')
            logo_img.add_header('Content-Disposition', 'inline', filename='logo_dolorosa.png')
            msg.attach(logo_img)
    msg.send(fail_silently=False)


def registro_view(request):
    if request.user.is_authenticated:
        return redirect('horarios:index')

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        username   = request.POST.get('username', '').strip()
        email      = request.POST.get('email', '').strip()
        rol        = request.POST.get('rol', 'estudiante')
        password   = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')

        # Validaciones
        if password != confirm_password:
            messages.error(request, 'Las contraseñas no coinciden.')
            return render(request, 'horarios/registro.html')

        if not email.endswith('@ladolorosa-loja.edu.ec'):
            messages.warning(request, 'El correo debe pertenecer a la institución (@ladolorosa-loja.edu.ec).')
            return render(request, 'horarios/registro.html')

        if Usuario.objects.filter(username=username).exists():
            messages.error(request, 'Este nombre de usuario ya está en uso.')
            return render(request, 'horarios/registro.html')

        if Usuario.objects.filter(email=email).exists():
            messages.error(request, 'Ya existe una cuenta con ese correo electrónico.')
            return render(request, 'horarios/registro.html')

        # Generar código y guardar datos en sesión (no crear cuenta aún)
        code = str(random.randint(100000, 999999))
        request.session['registro_pending'] = {
            'first_name': first_name,
            'last_name':  last_name,
            'username':   username,
            'email':      email,
            'rol':        rol,
            'password':   password,
            'code':       code,
            'expires':    (timezone.now() + timedelta(minutes=15)).isoformat(),
        }

        try:
            _send_registro_code(email, first_name, code)
        except Exception as e:
            messages.error(request, f'No se pudo enviar el correo de verificación: {str(e)}')
            return render(request, 'horarios/registro.html')

        return redirect('horarios:verify_registro')

    return render(request, 'horarios/registro.html')


def verify_registro_view(request):
    pending = request.session.get('registro_pending')
    if not pending:
        messages.error(request, 'No hay un registro pendiente. Por favor, completa el formulario primero.')
        return redirect('horarios:registro')

    email = pending.get('email', '')

    if request.method == 'POST':
        action = request.POST.get('action')

        # Reenviar código
        if action == 'resend':
            code = str(random.randint(100000, 999999))
            pending['code'] = code
            pending['expires'] = (timezone.now() + timedelta(minutes=15)).isoformat()
            request.session['registro_pending'] = pending
            request.session.modified = True
            try:
                _send_registro_code(email, pending.get('first_name', ''), code)
                messages.success(request, 'Se ha reenviado un nuevo código a tu correo.')
            except Exception as e:
                messages.error(request, f'Error al reenviar el código: {str(e)}')
            return redirect('horarios:verify_registro')

        # Verificar código
        code_input = request.POST.get('code', '').strip()
        stored_code = pending.get('code')
        expires_str = pending.get('expires')

        # Comprobar expiración
        from datetime import datetime
        try:
            expires = datetime.fromisoformat(expires_str).replace(tzinfo=timezone.utc)
        except Exception:
            expires = timezone.now() - timedelta(seconds=1)

        if code_input != stored_code:
            messages.error(request, 'Código incorrecto. Inténtalo de nuevo.')
            return render(request, 'horarios/verify_registro.html', {'email': email})

        if timezone.now() > expires:
            messages.error(request, 'El código ha expirado. Solicita uno nuevo.')
            return render(request, 'horarios/verify_registro.html', {'email': email})

        # Código válido → crear la cuenta
        try:
            user = Usuario.objects.create_user(
                username=pending['username'],
                email=pending['email'],
                password=pending['password'],
                first_name=pending['first_name'],
                last_name=pending['last_name'],
                rol=pending['rol'],
            )
            if pending['rol'] == 'docente':
                Docente.objects.create(usuario=user)

            # Limpiar sesión
            del request.session['registro_pending']

            messages.success(request, '¡Cuenta creada exitosamente! Ahora puedes iniciar sesión.')
            next_url = request.session.pop('registro_next', None)
            if next_url:
                return redirect(f"{reverse('horarios:login')}?next={next_url}")
            return redirect('horarios:login')
        except Exception as e:
            messages.error(request, f'Ocurrió un error al crear la cuenta: {str(e)}')

    return render(request, 'horarios/verify_registro.html', {'email': email})



def logout_view(request):
    logout(request)
    return redirect('horarios:login')

def verify_2fa_view(request):
    user_id = request.session.get('pre_2fa_user_id')
    if not user_id:
        return redirect('horarios:login')
        
    user = get_object_or_404(Usuario, id=user_id)
    
    if request.method == 'POST':
        code = request.POST.get('code')
        if user.two_factor_code == code and user.two_factor_expires and user.two_factor_expires > timezone.now():
            user.two_factor_code = None
            user.two_factor_expires = None
            user.save()
            
            login(request, user)
            del request.session['pre_2fa_user_id']
            
            if user.is_docente():
                return redirect('horarios:dashboard_docente')
            elif user.is_admin() or user.is_superuser:
                return redirect('horarios:gestion_dashboard')
            else:
                return redirect('horarios:index')
        else:
            messages.error(request, 'Código inválido o expirado.')
            
    return render(request, 'horarios/verify_2fa.html', {'user_email': user.email})

@login_required
def configuracion_view(request):
    user = request.user
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'change_password':
            password_form = PasswordChangeForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Tu contraseña ha sido actualizada exitosamente.')
                return redirect('horarios:configuracion')
            else:
                messages.error(request, 'Por favor corrige los errores abajo para cambiar tu contraseña.')
                
        elif action == 'toggle_2fa':
            password = request.POST.get('password')
            if user.check_password(password):
                user.two_factor_enabled = not user.two_factor_enabled
                user.save()
                estado = 'activada' if user.two_factor_enabled else 'desactivada'
                messages.success(request, f'La verificación de dos pasos ha sido {estado}.')
                return redirect('horarios:configuracion')
            else:
                messages.error(request, 'Contraseña incorrecta. No se pudo cambiar la configuración de 2FA.')
                password_form = PasswordChangeForm(user)
    else:
        password_form = PasswordChangeForm(user)
        
    return render(request, 'horarios/configuracion.html', {
        'password_form': password_form,
    })

from django.http import JsonResponse
import json

@login_required
def toggle_modo_oscuro(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            modo_oscuro = data.get('modo_oscuro', False)
            request.user.modo_oscuro = modo_oscuro
            request.user.save()
            return JsonResponse({'status': 'success', 'modo_oscuro': request.user.modo_oscuro})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@login_required
def index(request):
    # Directorio de Docentes y Disponibilidad
    # Mostrar todos los perfiles docentes (incluyendo admins que dan clases),
    # excepto la cuenta base del sistema 'admin'
    docentes = Docente.objects.exclude(
        usuario__username='admin'
    ).order_by('usuario__first_name')
    # Excluir el propio perfil si el usuario autenticado es docente
    if request.user.is_docente():
        docentes = docentes.exclude(usuario=request.user)
        
    hoy = datetime.datetime.now().weekday()
    dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
    dia_actual_str = dias_semana[hoy]
    
    hay_atencion_hoy = Horario.objects.filter(tipo='atencion', dia=dia_actual_str).exists()
        
    context = {
        'docentes': docentes,
        'hay_atencion_hoy': hay_atencion_hoy,
        'dia_actual_str': dia_actual_str.capitalize(),
    }
    return render(request, 'horarios/index.html', context)

@login_required
def dashboard_docente(request):
    if not (request.user.is_docente() or request.user.is_admin() or request.user.is_superuser):
        return redirect('horarios:index')
    
    # Crear un perfil de Docente temporal o permanente para el administrador si no lo tiene
    if not hasattr(request.user, 'perfil_docente'):
        Docente.objects.create(usuario=request.user, especialidad="Administrador")
    
    docente = request.user.perfil_docente

    if request.method == 'POST':
        form = DocenteFotoForm(request.POST, request.FILES, instance=docente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Foto de perfil actualizada exitosamente.')
            return redirect('horarios:dashboard_docente')
        else:
            messages.error(request, 'Error al subir la foto de perfil.')

    horarios = Horario.objects.filter(docente=docente).exclude(tipo='atencion').order_by('dia', 'hora_inicio')
    
    # Para el grid interactivo, fusionamos bloques y necesitamos celdas vacías
    horarios_grid, celdas_vacias = preparar_horarios_grid(horarios, incluir_vacios=True, fusionar_bloques=True, color_by='default')
    
    context = {
        'docente': docente,
        'horarios': horarios,
        'horarios_grid': horarios_grid,
        'celdas_vacias': celdas_vacias,
        'cursos': Curso.objects.all(),
        'asignaturas': Asignatura.objects.all(),
        'dias': Horario.DIAS,
        'es_docentes': True,
        'paralelos_json': _paralelos_json(),
    }
    return render(request, 'horarios/dashboard_docente.html', context)

@login_required
def imprimir_horario(request):
    docente_id = request.GET.get('docente_id')
    curso_id = request.GET.get('curso_id')
    paralelo_id = request.GET.get('paralelo_id')

    es_clase = False
    titulo_horario = ""
    tutor_nombre = ""

    if curso_id and paralelo_id:
        es_clase = True
        curso = get_object_or_404(Curso, id=curso_id)
        paralelo = get_object_or_404(Paralelo, id=paralelo_id)
        horarios = Horario.objects.filter(curso=curso, paralelo=paralelo).exclude(tipo='atencion').select_related('asignatura', 'docente__usuario').order_by('dia', 'hora_inicio')
        titulo_horario = f"{curso.nombre} '{paralelo.identificador}'"
        if paralelo.tutor:
            tutor_nombre = paralelo.tutor.usuario.get_full_name() or paralelo.tutor.usuario.username
    elif docente_id:
        docente = get_object_or_404(Docente, id=docente_id)
        horarios = Horario.objects.filter(docente=docente).exclude(tipo='atencion').select_related('asignatura', 'curso', 'paralelo').order_by('dia', 'hora_inicio')
        titulo_horario = docente.usuario.get_full_name() or docente.usuario.username
    else:
        if hasattr(request.user, 'perfil_docente'):
            docente = request.user.perfil_docente
            horarios = Horario.objects.filter(docente=docente).exclude(tipo='atencion').select_related('asignatura', 'curso', 'paralelo').order_by('dia', 'hora_inicio')
            titulo_horario = docente.usuario.get_full_name() or docente.usuario.username
        else:
            messages.error(request, 'No tienes un perfil de docente y no seleccionaste un horario específico.')
            return redirect('horarios:index')

    tabla = generar_tabla_impresion(horarios, es_clase)

    from django.utils import timezone
    fecha_hoy = timezone.localdate()

    context = {
        'es_clase': es_clase,
        'titulo_horario': titulo_horario,
        'tutor_nombre': tutor_nombre,
        'tabla': tabla,
        'fecha_hoy': fecha_hoy,
    }
    return render(request, 'horarios/imprimir_horario.html', context)

@login_required
@require_POST
def cambiar_estado(request):
    if not hasattr(request.user, 'perfil_docente'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    
    try:
        data = json.loads(request.body)
        docente = request.user.perfil_docente
        docente.disponible = data.get('disponible', docente.disponible)
        docente.ubicacion_actual = data.get('ubicacion_actual', docente.ubicacion_actual)
        docente.save()
        return JsonResponse({'status': 'ok', 'disponible': docente.estado_disponible, 'ubicacion': docente.ubicacion_dinamica})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def horarios_curso(request):
    cursos = Curso.objects.all()
    # Identificadores únicos de paralelos (A, B, C...) en lugar de cada paralelo
    identificadores = Paralelo.objects.values_list('identificador', flat=True).distinct().order_by('identificador')
    
    curso_id = request.GET.get('curso')
    paralelo_ident = request.GET.get('paralelo')
    
    grupos = []
    
    paralelos = Paralelo.objects.select_related('curso').all()
    if curso_id:
        paralelos = paralelos.filter(curso_id=curso_id)
    if paralelo_ident:
        paralelos = paralelos.filter(identificador=paralelo_ident)
        
    # Mostrar todos los horarios (docente y clase) en la vista del curso,
    # ya que la validación de choques previene duplicados.
    horarios_qs = Horario.objects.select_related('docente__usuario', 'asignatura')
    
    for p in paralelos:
        hs = horarios_qs.filter(curso=p.curso, paralelo=p).order_by('dia', 'hora_inicio')
        # Sólo mostrar grupos que tienen algún horario para evitar listas larguísimas vacías
        if hs.exists() or (curso_id or paralelo_ident): 
            horarios_grid = preparar_horarios_grid(hs, fusionar_bloques=True, color_by='asignatura')
            
            grupos.append({
                'id': f"curso_{p.curso_id}_paralelo_{p.id}",
                'curso_id': p.curso_id,
                'paralelo_id': p.id,
                'titulo': f"{p.curso.nombre} '{p.identificador}'",
                'horarios_grid': horarios_grid,
            })
            
    grupos.sort(key=lambda x: (_get_curso_order(x['titulo']), x['titulo']))
        
    context = {
        'cursos': cursos,
        'identificadores': identificadores,
        'grupos': grupos,
    }
    return render(request, 'horarios/horarios_curso.html', context)

@login_required
def horario_docente(request, docente_id):
    """Vista pública (para docentes y estudiantes) del horario de un docente específico."""
    docente = get_object_or_404(Docente, pk=docente_id)
    horarios = Horario.objects.filter(
        docente=docente
    ).exclude(tipo='atencion').order_by('dia', 'hora_inicio')
    
    asignaturas_count = horarios.values('asignatura').distinct().count()
    cursos_count = horarios.values('curso', 'paralelo').distinct().count()
    
    horarios_grid = preparar_horarios_grid(horarios, fusionar_bloques=True, color_by='default')

    context = {
        'docente': docente,
        'horarios': horarios,
        'horarios_grid': horarios_grid,
        'asignaturas_count': asignaturas_count,
        'cursos_count': cursos_count,
        'es_propio': request.user.is_docente() and hasattr(request.user, 'perfil_docente') and request.user.perfil_docente == docente,
    }
    return render(request, 'horarios/horario_docente.html', context)

@login_required
def añadir_horario(request):
    if not request.user.is_docente() and not request.user.is_admin():
        messages.error(request, 'No tienes permisos para añadir horarios.')
        return redirect('horarios:index')

    if request.method == 'POST':
        form = HorarioForm(request.POST)
        if form.is_valid():
            horario = form.save(commit=False)
            horario.docente = request.user.perfil_docente
            horario.save()
            messages.success(request, 'Horario añadido exitosamente.')
            return redirect('horarios:dashboard_docente')
    else:
        form = HorarioForm()

    # Construir mapa curso_id -> lista de paralelos para filtrado en JS
    paralelos_por_curso = {}
    for p in Paralelo.objects.select_related('curso').all():
        cid = str(p.curso_id)
        if cid not in paralelos_por_curso:
            paralelos_por_curso[cid] = []
        label = p.identificador
        if p.especialidad_bachillerato:
            label = f"{p.identificador} ({p.get_especialidad_bachillerato_display()})"
        paralelos_por_curso[cid].append({'id': p.id, 'nombre': label})

    return render(request, 'horarios/anadir_horario.html', {
        'form': form,
        'paralelos_json': json.dumps(paralelos_por_curso),
    })


# ──────────────────────────────────────────────────────────────────────────────
# PANEL DE GESTIÓN (solo administradores)
# ──────────────────────────────────────────────────────────────────────────────

def _paralelos_json():
    """Construye el mapa curso_id → paralelos para el JS de cascada."""
    mapa = {}
    for p in Paralelo.objects.select_related('curso').all():
        cid = str(p.curso_id)
        if cid not in mapa:
            mapa[cid] = []
        label = p.identificador
        if p.especialidad_bachillerato:
            label = f"{p.identificador} ({p.get_especialidad_bachillerato_display()})"
        mapa[cid].append({'id': p.id, 'nombre': label})
    return json.dumps(mapa)


def _todos_paralelos_json():
    """Lista plana de todos los paralelos con metadata del curso para JS."""
    datos = []
    for p in Paralelo.objects.select_related('curso').all().order_by('curso__nombre', 'identificador'):
        datos.append({
            'id': p.id,
            'identificador': p.identificador,
            'curso_nombre': p.curso.nombre,
            'curso_nombre_lower': p.curso.nombre.lower(),
            'especialidad': p.especialidad_bachillerato or '',
            'label': f"{p.curso.nombre} – {p.identificador}"
        })
    return json.dumps(datos)


@admin_required
def gestion_dashboard(request):
    context = {
        'total_docentes':        Docente.objects.count(),
        'total_cursos':          Curso.objects.count(),
        'total_paralelos':       Paralelo.objects.count(),
        'total_h_docente':       Horario.objects.filter(tipo='docente').count(),
        'total_h_clase':         Horario.objects.filter(tipo='clase').count(),
        'docentes_disponibles':  Docente.objects.filter(disponible=True).count(),
        'ultimos_horarios':      Horario.objects.select_related(
                                    'docente__usuario', 'curso', 'paralelo'
                                 ).order_by('-id')[:8],
    }
    return render(request, 'horarios/gestion/dashboard.html', context)


@admin_required
def gestion_horarios(request):
    return redirect('horarios:gestion_horarios_docentes')

@admin_required
def gestion_horarios_docentes(request):
    return _render_horarios_list(request, 'docente', 'Horarios de Docentes')

@admin_required
def gestion_horarios_cursos(request):
    return _render_horarios_list(request, 'clase', 'Horarios de Cursos')

@admin_required
def gestion_horarios_preliminar(request):
    return _render_horarios_list(request, 'clase', 'Vista Preliminar de Horarios', template_name='horarios/gestion/horarios_preliminar.html')

def _get_curso_order(titulo):
    t = titulo.lower()
    if 'octavo' in t: return 1
    if 'noveno' in t: return 2
    if 'décimo' in t or 'decimo' in t: return 3
    if 'primero' in t: return 4
    if 'segundo' in t: return 5
    if 'tercero' in t: return 6
    return 99

def generar_tabla_impresion(horarios, es_clase=True):
    PERIODOS = [
        {'num': '1',  'inicio': '07:00', 'fin': '07:45', 'key': '07:00:00'},
        {'num': '2',  'inicio': '07:45', 'fin': '08:30', 'key': '07:45:00'},
        {'num': '3',  'inicio': '08:30', 'fin': '09:15', 'key': '08:30:00'},
        {'num': '4',  'inicio': '09:15', 'fin': '10:00', 'key': '09:15:00'},
        {'num': 'R',  'inicio': '10:00', 'fin': '10:25', 'key': '10:00:00', 'recreo': True},
        {'num': '5',  'inicio': '10:25', 'fin': '11:05', 'key': '10:25:00'},
        {'num': '6',  'inicio': '11:05', 'fin': '11:45', 'key': '11:05:00'},
        {'num': '7',  'inicio': '11:45', 'fin': '12:25', 'key': '11:45:00'},
        {'num': '8',  'inicio': '12:25', 'fin': '13:05', 'key': '12:25:00'},
        {'num': '9',  'inicio': '13:05', 'fin': '13:45', 'key': '13:05:00'},
    ]
    DIAS = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']

    horario_map = {}
    for h in horarios:
        key_dia = h.dia.lower()
        key_hora = h.hora_inicio.strftime("%H:%M:%S")
        horario_map[(key_dia, key_hora)] = h

    def color_para(h):
        if es_clase:
            clave = f"{h.asignatura.nombre if h.asignatura else ''}|{h.docente_id}"
        else:
            clave = f"{h.asignatura.nombre if h.asignatura else ''}|{h.curso_id}|{h.paralelo_id}"
        return get_color_for_materia(clave)

    tabla = []
    skip_map = {}
    for row_idx, periodo in enumerate(PERIODOS):
        if periodo.get('recreo'):
            tabla.append({'periodo': periodo, 'recreo': True})
            continue
            
        fila = {'periodo': periodo, 'recreo': False, 'dias': []}
        for dia_idx, dia in enumerate(DIAS):
            if skip_map.get((dia_idx, row_idx)):
                continue
                
            h = horario_map.get((dia, periodo['key']))
            if h:
                rowspan = 1
                for next_row_idx in range(row_idx + 1, len(PERIODOS)):
                    next_periodo = PERIODOS[next_row_idx]
                    if next_periodo.get('recreo'):
                        break
                    next_h = horario_map.get((dia, next_periodo['key']))
                    if next_h:
                        if es_clase:
                            mismo_bloque = (next_h.asignatura_id == h.asignatura_id and next_h.docente_id == h.docente_id)
                        else:
                            mismo_bloque = (next_h.asignatura_id == h.asignatura_id and next_h.curso_id == h.curso_id and next_h.paralelo_id == h.paralelo_id)
                        
                        if mismo_bloque:
                            rowspan += 1
                            skip_map[(dia_idx, next_row_idx)] = True
                        else:
                            break
                    else:
                        break
                        
                colors = color_para(h)
                fila['dias'].append({
                    'horario': h,
                    'color_bg': colors['bg'],
                    'color_border': colors['border'],
                    'color_text': colors['text'],
                    'rowspan': rowspan,
                })
            else:
                fila['dias'].append(None)
        tabla.append(fila)
    return tabla

def _render_horarios_list(request, tipo_agrupacion, titulo, template_name='horarios/gestion/horarios.html'):
    horarios_qs = Horario.objects.select_related('docente__usuario', 'curso', 'paralelo').exclude(tipo='atencion')
    atencion_qs = Horario.objects.select_related('docente__usuario', 'curso', 'paralelo', 'asignatura').filter(tipo='atencion')
    
    docente_id = request.GET.get('docente')
    curso_id   = request.GET.get('curso')
    dia        = request.GET.get('dia')

    if docente_id: 
        horarios_qs = horarios_qs.filter(docente_id=docente_id)
        atencion_qs = atencion_qs.filter(docente_id=docente_id)
    if curso_id:   
        horarios_qs = horarios_qs.filter(curso_id=curso_id)
        atencion_qs = atencion_qs.filter(curso_id=curso_id)
    if dia:        
        horarios_qs = horarios_qs.filter(dia=dia)
        atencion_qs = atencion_qs.filter(dia=dia)

    grupos = []
    
    if tipo_agrupacion == 'clase':
        paralelos = Paralelo.objects.select_related('curso').all()
        if curso_id:
            paralelos = paralelos.filter(curso_id=curso_id)
        
        for p in paralelos:
            hs = horarios_qs.filter(curso=p.curso, paralelo=p).order_by('dia', 'hora_inicio')
            atencion_hs = atencion_qs.filter(curso=p.curso, paralelo=p).order_by('dia', 'hora_inicio')
            filled, empty = preparar_horarios_grid(hs, incluir_vacios=True, fusionar_bloques=True, color_by='asignatura')
            tutor_nombre = ""
            if p.tutor:
                tutor_nombre = p.tutor.usuario.get_full_name() or p.tutor.usuario.username
            grupos.append({
                'id': f"curso_{p.curso_id}_paralelo_{p.id}",
                'titulo': f"{p.curso.nombre} '{p.identificador}'",
                'curso_id': p.curso_id,
                'paralelo_id': p.id,
                'tutor_nombre': tutor_nombre,
                'horarios_grid': filled,
                'celdas_vacias': empty,
                'horarios_atencion': atencion_hs,
                'tabla': generar_tabla_impresion(hs, True),
            })
        # Ordenar cronológicamente (Octavo -> Tercero) y luego por identificador
        grupos.sort(key=lambda x: (_get_curso_order(x['titulo']), x['titulo']))
    else:
        docentes_qs = Docente.objects.select_related('usuario').all().order_by('usuario__first_name')
        if docente_id:
            docentes_qs = docentes_qs.filter(id=docente_id)
            
        for d in docentes_qs:
            hs = horarios_qs.filter(docente=d).order_by('dia', 'hora_inicio')
            atencion_hs = atencion_qs.filter(docente=d).order_by('dia', 'hora_inicio')
            filled, empty = preparar_horarios_grid(hs, incluir_vacios=True, fusionar_bloques=True, color_by='default')
            grupos.append({
                'id': f"docente_{d.id}",
                'titulo': d.usuario.get_full_name() or d.usuario.username,
                'docente_id': d.id,
                'horarios_grid': filled,
                'celdas_vacias': empty,
                'horarios_atencion': atencion_hs,
                'tabla': generar_tabla_impresion(hs, False),
            })

    context = {
        'grupos':    grupos,
        'docentes':  Docente.objects.select_related('usuario').order_by('usuario__first_name'),
        'cursos':    Curso.objects.all(),
        'dias':      Horario.DIAS,
        'filtros':   {'docente': docente_id, 'curso': curso_id, 'dia': dia},
        'titulo_seccion': titulo,
        'es_docentes': tipo_agrupacion == 'docente',
        'hay_resultados': len(grupos) > 0,
        'paralelos_json': _paralelos_json(),
        'asignaturas': Asignatura.objects.all(),
    }
    return render(request, template_name, context)

@admin_required
def exportar_horarios_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    import datetime
    
    tipo_agrupacion = 'clase' if not request.GET.get('docente') else 'docente'
    
    horarios_qs = Horario.objects.select_related('docente__usuario', 'curso', 'paralelo', 'asignatura').exclude(tipo='atencion')
    docente_id = request.GET.get('docente')
    curso_id   = request.GET.get('curso')
    dia        = request.GET.get('dia')

    if docente_id: horarios_qs = horarios_qs.filter(docente_id=docente_id)
    if curso_id:   horarios_qs = horarios_qs.filter(curso_id=curso_id)
    if dia:        horarios_qs = horarios_qs.filter(dia=dia)

    grupos = []
    if tipo_agrupacion == 'clase':
        paralelos = Paralelo.objects.select_related('curso').all()
        if curso_id:
            paralelos = paralelos.filter(curso_id=curso_id)
        for p in paralelos:
            hs = horarios_qs.filter(curso=p.curso, paralelo=p).order_by('dia', 'hora_inicio')
            if hs.exists():
                tabla = generar_tabla_impresion(hs, True)
                grupos.append({
                    'titulo': f"{p.curso.nombre} '{p.identificador}'",
                    'tabla': tabla
                })
        grupos.sort(key=lambda x: (_get_curso_order(x['titulo']), x['titulo']))
    else:
        docentes_qs = Docente.objects.select_related('usuario').all().order_by('usuario__first_name')
        if docente_id:
            docentes_qs = docentes_qs.filter(id=docente_id)
        for d in docentes_qs:
            hs = horarios_qs.filter(docente=d).order_by('dia', 'hora_inicio')
            if hs.exists():
                tabla = generar_tabla_impresion(hs, False)
                grupos.append({
                    'titulo': d.usuario.get_full_name() or d.usuario.username,
                    'tabla': tabla
                })

    wb = Workbook()
    ws = wb.active
    ws.title = "Horarios"
    
    # Configurar página para Landscape, encajar todas las columnas en 1 página de ancho
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0 

    # Estilos exactos
    font_main_title = Font(name='Times New Roman', size=13, bold=True, color="FF1A237E", italic=True)
    font_motto = Font(name='Arial', size=9, color="FF5C6BC0", italic=True)
    font_pill = Font(name='Arial', size=11, bold=True, color="FFFFFFFF")
    fill_pill = PatternFill(start_color="FF7986CB", end_color="FF7986CB", fill_type="solid")
    
    font_table_head = Font(name='Arial', size=8, bold=True, color="FF1A237E")
    fill_table_head = PatternFill(start_color="FFF8F9FA", end_color="FFF8F9FA", fill_type="solid")
    
    font_time_num = Font(name='Arial', size=7, bold=True, color="FF3949AB")
    font_time_hours = Font(name='Arial', size=6, color="FF555555")
    
    font_subject = Font(name='Arial', size=8, color="FF111111")
    font_teacher = Font(name='Arial', size=7, color="FF333333", italic=True)
    
    font_recreo = Font(name='Arial', size=8, bold=True, color="FF283593")
    fill_recreo = PatternFill(start_color="FFE8EAF6", end_color="FFE8EAF6", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color="FF9FA8DA"), 
        right=Side(style='thin', color="FF9FA8DA"), 
        top=Side(style='thin', color="FF9FA8DA"), 
        bottom=Side(style='thin', color="FF9FA8DA")
    )
    thick_border = Border(
        left=Side(style='medium', color="FF283593"), 
        right=Side(style='medium', color="FF283593"), 
        top=Side(style='medium', color="FF283593"), 
        bottom=Side(style='medium', color="FF283593")
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')

    dias_nombres = {0: 'Lu', 1: 'Ma', 2: 'Mi', 3: 'Ju', 4: 'Vi'}
    
    def render_table_excel(ws, grupo, start_row, base_col):
        # HEADER (3 rows)
        # Row 1: Title
        ws.merge_cells(start_row=start_row, start_column=base_col, end_row=start_row, end_column=base_col+5)
        c_title = ws.cell(row=start_row, column=base_col, value="Unidad Educativa Fiscomisional La Dolorosa")
        c_title.font = font_main_title
        c_title.alignment = align_center
        
        # Row 2: Motto & Pill
        ws.merge_cells(start_row=start_row+1, start_column=base_col, end_row=start_row+1, end_column=base_col+2)
        c_motto = ws.cell(row=start_row+1, column=base_col, value="¡Siempre! ... un paso adelante")
        c_motto.font = font_motto
        c_motto.alignment = align_left
        
        ws.merge_cells(start_row=start_row+1, start_column=base_col+3, end_row=start_row+1, end_column=base_col+5)
        c_pill = ws.cell(row=start_row+1, column=base_col+3, value=grupo['titulo'])
        c_pill.font = font_pill
        c_pill.fill = fill_pill
        c_pill.alignment = align_center
        
        grid_start_row = start_row + 3
        
        # TABLE HEAD
        c_hora = ws.cell(row=grid_start_row, column=base_col, value="")
        c_hora.fill = fill_table_head
        c_hora.border = thick_border
        
        for col_idx in range(0, 5):
            c_dia = ws.cell(row=grid_start_row, column=base_col + 1 + col_idx, value=dias_nombres[col_idx])
            c_dia.font = font_table_head
            c_dia.alignment = align_center
            c_dia.fill = fill_table_head
            c_dia.border = thick_border
            
        current_row = grid_start_row + 1
        skip_cells = {} # To handle rowspans manually
        
        for fila_idx, fila in enumerate(grupo['tabla']):
            excel_r = current_row
            
            # Hora col
            cell_hora = ws.cell(row=excel_r, column=base_col)
            cell_hora.alignment = align_center
            cell_hora.border = thin_border
            cell_hora.fill = fill_table_head
            
            if fila.get('recreo'):
                cell_hora.value = "REC\n10:00 - 10:25"
                cell_hora.font = font_time_num
                
                ws.merge_cells(start_row=excel_r, start_column=base_col+1, end_row=excel_r, end_column=base_col+5)
                c_rec = ws.cell(row=excel_r, column=base_col+1, value="Recreo")
                c_rec.alignment = align_center
                c_rec.font = font_recreo
                c_rec.fill = fill_recreo
                for c_idx in range(1, 6):
                    ws.cell(row=excel_r, column=base_col + c_idx).border = thin_border
            else:
                periodo = fila['periodo']
                cell_hora.value = f"{periodo['num']}\n{periodo['inicio'][:5]} - {periodo['fin'][:5]}"
                cell_hora.font = font_time_num
                
                for dia_idx, celda in enumerate(fila.get('dias', [])):
                    col_offset = dia_idx + 1
                    excel_c = base_col + col_offset
                    
                    if skip_cells.get((excel_r, excel_c)):
                        continue
                        
                    if celda:
                        h = celda['horario']
                        rowspan = celda.get('rowspan', 1)
                        
                        texto = h.asignatura.nombre if getattr(h, 'asignatura', None) else getattr(h, 'materia', '')
                        subtexto = h.docente.usuario.get_full_name() if tipo_agrupacion == 'clase' else f"{h.curso.nombre} '{h.paralelo.identificador}'"
                        valor = f"{texto}\n{subtexto}"
                        
                        excel_r_end = excel_r + rowspan - 1
                        
                        if rowspan > 1:
                            ws.merge_cells(start_row=excel_r, start_column=excel_c, end_row=excel_r_end, end_column=excel_c)
                            for r_skip in range(excel_r + 1, excel_r_end + 1):
                                skip_cells[(r_skip, excel_c)] = True
                                
                        cell_d = ws.cell(row=excel_r, column=excel_c, value=valor)
                        cell_d.alignment = align_center
                        cell_d.font = font_subject
                        
                        color_bg = celda.get('color_bg', '#ffffff').replace('#', '').strip()
                        if len(color_bg) == 3: color_bg = color_bg[0]*2 + color_bg[1]*2 + color_bg[2]*2
                        if len(color_bg) == 6: color_bg = "FF" + color_bg
                        color_bg = color_bg.upper()
                        if len(color_bg) != 8: color_bg = "FFFFFFFF"
                        
                        cell_d.fill = PatternFill(start_color=color_bg, end_color=color_bg, fill_type="solid")
                        
                        for merge_r in range(excel_r, excel_r_end + 1):
                            ws.cell(row=merge_r, column=excel_c).border = thin_border
                    else:
                        c_vacia = ws.cell(row=excel_r, column=excel_c, value="")
                        c_vacia.border = thin_border
                        
            current_row += 1

        # Footer
        footer_row = current_row
        ws.merge_cells(start_row=footer_row, start_column=base_col, end_row=footer_row, end_column=base_col+5)
        c_foot = ws.cell(row=footer_row, column=base_col, value="Elaborado por JRSZ" if tipo_agrupacion == 'clase' else "Horario Docente elaborado por JRSZ")
        c_foot.font = font_motto
        c_foot.alignment = align_right

    row_cursor = 1
    # Dibujar grupos en pares de 2 (Grid 2 columnas)
    for i in range(0, len(grupos), 2):
        grupo1 = grupos[i]
        grupo2 = grupos[i+1] if i+1 < len(grupos) else None
        
        render_table_excel(ws, grupo1, row_cursor, 1)
        if grupo2:
            render_table_excel(ws, grupo2, row_cursor, 8)
            
        row_cursor += 17 # Espacio para la siguiente fila de 2

    # Ajustar ancho columnas
    for c_base in [1, 8]:
        ws.column_dimensions[chr(64 + c_base)].width = 11 # Hora Col (A, H)
        for c_offset in range(1, 6):
            ws.column_dimensions[chr(64 + c_base + c_offset)].width = 16 # Dias (B-F, I-M)
    
    ws.column_dimensions['G'].width = 3 # Columna vacía separadora

    # Fix for height
    for r in range(1, row_cursor):
        ws.row_dimensions[r].height = 28 # Make rows taller to fit subject and teacher comfortably

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=horarios_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

@login_required
@require_POST
def api_guardar_horario(request):
    horario_id = request.POST.get('horario_id')
    dia = request.POST.get('dia')
    hora_inicio = request.POST.get('hora_inicio')
    hora_fin = request.POST.get('hora_fin')
    asignatura_id = request.POST.get('asignatura_id')
    docente_id = request.POST.get('docente_id')
    curso_id = request.POST.get('curso_id')
    paralelo_id = request.POST.get('paralelo_id')

    if not all([dia, hora_inicio, hora_fin, asignatura_id, docente_id, curso_id, paralelo_id]):
        return JsonResponse({'status': 'error', 'message': "Faltan datos requeridos para guardar el horario."})

    # Permisos
    if not (request.user.is_admin() or request.user.is_superuser):
        if not request.user.is_docente():
            return JsonResponse({'status': 'error', 'message': "No tienes permisos."}, status=403)
        if str(request.user.perfil_docente.id) != str(docente_id):
            return JsonResponse({'status': 'error', 'message': "Solo puedes editar tus propios horarios."}, status=403)

    try:
        with transaction.atomic():
            docente = Docente.objects.get(id=docente_id)
            curso = Curso.objects.get(id=curso_id)
        paralelo = Paralelo.objects.get(id=paralelo_id)
        asignatura = Asignatura.objects.get(id=asignatura_id)
        
        cantidad_periodos = int(request.POST.get('cantidad_periodos', 1))
        
        periodos_a_guardar = []
        hora_inicio_str = str(hora_inicio)
        if len(hora_inicio_str) == 5:
            hora_inicio_str += ":00"
            
        start_idx = -1
        for i, row in enumerate(PERIODOS_ROWS):
            if row[1] == hora_inicio_str:
                start_idx = i
                break
                
        if start_idx != -1:
            for i in range(cantidad_periodos):
                if start_idx + i < len(PERIODOS_ROWS):
                    periodos_a_guardar.append((PERIODOS_ROWS[start_idx + i][1], PERIODOS_ROWS[start_idx + i][2]))
                else:
                    break
                    
        if not periodos_a_guardar:
            periodos_a_guardar = [(hora_inicio, hora_fin)]

        # Validación de cruce de horarios para todos los periodos
        # Primero recopilamos los IDs de los bloques que vamos a sobrescribir (si es un edit de un bloque fusionado)
        ids_a_excluir = []
        if horario_id and horario_id != "null" and horario_id != "":
            h_base = Horario.objects.get(id=horario_id)
            ids_a_excluir.append(h_base.id)
            # Buscar otros bloques continuos de la misma clase que conforman el bloque fusionado actual
            # para excluirlos todos de la validación de choque y luego sobrescribirlos/eliminarlos.
            bloques_fusionados = Horario.objects.filter(
                docente=h_base.docente, curso=h_base.curso, paralelo=h_base.paralelo,
                asignatura=h_base.asignatura, dia=h_base.dia
            ).exclude(id=h_base.id)
            
            # Simple heurística: Si un bloque adyacente tiene la misma materia, asumimos que era parte del bloque fusionado
            # En la práctica, es más seguro eliminar TODOS los bloques de esta materia continuos e insertar los nuevos
            for bf in bloques_fusionados:
                ids_a_excluir.append(bf.id)
        
        for p_inicio, p_fin in periodos_a_guardar:
            # Check docente
            choques_docente = Horario.objects.filter(
                docente=docente,
                dia=dia,
                hora_inicio__lt=p_fin,
                hora_fin__gt=p_inicio
            )
            if ids_a_excluir:
                choques_docente = choques_docente.exclude(id__in=ids_a_excluir)
                
            if choques_docente.exists():
                choque = choques_docente.first()
                docente_nombre = docente.usuario.get_full_name() or docente.usuario.username
                hora_i = choque.hora_inicio.strftime('%H:%M')
                hora_f = choque.hora_fin.strftime('%H:%M')
                return JsonResponse({'status': 'error', 'message': f"¡Choque de Horario! El docente {docente_nombre} ya dicta '{choque.asignatura.nombre}' el {dia} de {hora_i} a {hora_f}."})

            # Check curso/paralelo
            choques_curso = Horario.objects.filter(
                curso=curso,
                paralelo=paralelo,
                dia=dia,
                hora_inicio__lt=p_fin,
                hora_fin__gt=p_inicio
            )
            if ids_a_excluir:
                choques_curso = choques_curso.exclude(id__in=ids_a_excluir)
                
            if choques_curso.exists():
                choque = choques_curso.first()
                doc_choque_nombre = choque.docente.usuario.get_full_name() or choque.docente.usuario.username
                hora_i = choque.hora_inicio.strftime('%H:%M')
                hora_f = choque.hora_fin.strftime('%H:%M')
                return JsonResponse({'status': 'error', 'message': f"¡Choque de Horario! El curso {curso.nombre} '{paralelo.identificador}' ya tiene la clase '{choque.asignatura.nombre}' con el docente {doc_choque_nombre} el {dia} de {hora_i} a {hora_f}."})
        
        # Eliminar los bloques anteriores si estamos editando
        if ids_a_excluir:
            Horario.objects.filter(id__in=ids_a_excluir).delete()
        
        first_h = None
        tipo = 'docente' if request.POST.get('es_docentes') == '1' else 'clase'
        
        saved_horarios = []
        for idx, (p_inicio, p_fin) in enumerate(periodos_a_guardar):
            h = Horario.objects.create(
                dia=dia,
                hora_inicio=p_inicio,
                hora_fin=p_fin,
                asignatura=asignatura,
                docente=docente,
                curso=curso,
                paralelo=paralelo,
                tipo=tipo
            )
            saved_horarios.append(h)

        msg = f"Se guardaron {len(periodos_a_guardar)} periodos exitosamente." if len(periodos_a_guardar) > 1 else ("Horario actualizado exitosamente." if horario_id else "Horario guardado exitosamente.")

        es_docentes = request.POST.get('es_docentes') == '1'

        # Fetch ALL horarios for the affected group so the grid can be fully rebuilt client-side
        if es_docentes:
            grupo_id = f"docente_{docente.id}"
            all_horarios = Horario.objects.filter(docente=docente).select_related('docente__usuario', 'curso', 'paralelo', 'asignatura').order_by('dia', 'hora_inicio')
        else:
            grupo_id = f"curso_{curso.id}_paralelo_{paralelo.id}"
            all_horarios = Horario.objects.filter(curso=curso, paralelo=paralelo).select_related('docente__usuario', 'curso', 'paralelo', 'asignatura').order_by('dia', 'hora_inicio')

        from .views import preparar_horarios_grid
        all_filled = preparar_horarios_grid(all_horarios, fusionar_bloques=True, incluir_vacios=False)

        items_html_parts = []
        for item_data in all_filled:
            items_html_parts.append(render_to_string('horarios/gestion/schedule_item.html', {'item': item_data, 'es_docentes': es_docentes}))

        return JsonResponse({
            'status': 'success',
            'message': msg,
            'grupo_id': grupo_id,
            'items_html': "\n".join(items_html_parts),
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f"Error al guardar horario: {str(e)}"})          
@admin_required
def gestion_crear_horario(request):
    if request.method == 'POST':
        form = HorarioAdminForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario creado exitosamente.')
            return redirect('horarios:gestion_horarios')
    else:
        form = HorarioAdminForm()
    return render(request, 'horarios/gestion/horario_form.html', {
        'form': form,
        'paralelos_json': _paralelos_json(),
        'titulo': 'Crear Horario',
        'accion': 'Crear',
    })


@admin_required
def gestion_editar_horario(request, horario_id):
    horario = get_object_or_404(Horario, pk=horario_id)
    if request.method == 'POST':
        form = HorarioAdminForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario actualizado exitosamente.')
            return redirect('horarios:gestion_horarios')
    else:
        form = HorarioAdminForm(instance=horario)
    return render(request, 'horarios/gestion/horario_form.html', {
        'form': form,
        'paralelos_json': _paralelos_json(),
        'titulo': f'Editar — {horario.asignatura.nombre if horario.asignatura else "Horario"}',
        'accion': 'Guardar cambios',
        'horario': horario,
    })


@login_required
@require_POST
def gestion_eliminar_horario(request, horario_id):
    horario = get_object_or_404(Horario, pk=horario_id)
    
    # Permisos
    if not (request.user.is_admin() or request.user.is_superuser):
        if not request.user.is_docente() or horario.docente_id != request.user.perfil_docente.id:
            return JsonResponse({'status': 'error', 'message': "No autorizado para eliminar este horario."}, status=403)
            
    nombre = horario.asignatura.nombre if horario.asignatura else 'Horario'
    
    # Para eliminar el bloque completo (que visualmente estaba fusionado)
    horarios_to_delete = [horario]
    current_h = horario
    
    while True:
        # Buscamos el siguiente periodo que empiece exactamente donde termina este
        # y tenga exactamente los mismos atributos
        next_h = Horario.objects.filter(
            docente_id=current_h.docente_id,
            curso_id=current_h.curso_id,
            paralelo_id=current_h.paralelo_id,
            asignatura_id=current_h.asignatura_id,
            dia=current_h.dia,
            hora_inicio=current_h.hora_fin
        ).first()
        
        if next_h:
            horarios_to_delete.append(next_h)
            current_h = next_h
        else:
            break
            
    count = len(horarios_to_delete)
    for h in horarios_to_delete:
        h.delete()
        
    msg = f'Horario "{nombre}" eliminado correctamente.' if count == 1 else f'Bloque de {count} periodos de "{nombre}" eliminado correctamente.'
    return JsonResponse({'status': 'success', 'message': msg})


@admin_required
def gestion_docentes(request):
    docentes = Docente.objects.select_related('usuario').order_by('usuario__first_name')
    return render(request, 'horarios/gestion/docentes.html', {'docentes': docentes})

@admin_required
def gestion_crear_docente(request):
    if request.method == 'POST':
        form = DocenteCreateForm(request.POST)
        if form.is_valid():
            try:
                user = Usuario.objects.create_user(
                    username=form.cleaned_data['username'],
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    rol='docente'
                )
                docente = form.save(commit=False)
                docente.usuario = user
                docente.save()
                messages.success(request, 'Docente creado exitosamente.')
                return redirect('horarios:gestion_docentes')
            except Exception as e:
                messages.error(request, f'Ocurrió un error al crear el docente: {str(e)}')
    else:
        form = DocenteCreateForm()
    return render(request, 'horarios/gestion/docente_form.html', {
        'form': form, 'titulo': 'Crear Docente', 'accion': 'Crear'
    })

@admin_required
def gestion_editar_docente(request, docente_id):
    docente = get_object_or_404(Docente, pk=docente_id)
    if request.method == 'POST':
        form = DocenteEditForm(request.POST, instance=docente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Docente actualizado exitosamente.')
            return redirect('horarios:gestion_docentes')
    else:
        form = DocenteEditForm(instance=docente)
    return render(request, 'horarios/gestion/docente_form.html', {
        'form': form, 'docente': docente, 'titulo': f'Editar Docente', 'accion': 'Guardar cambios'
    })

@admin_required
@require_POST
def gestion_eliminar_docente(request, docente_id):
    docente = get_object_or_404(Docente, pk=docente_id)
    usuario = docente.usuario
    nombre = usuario.get_full_name() or usuario.username
    usuario.delete() # Al eliminar usuario, el docente se elimina en cascada
    messages.success(request, f'Docente "{nombre}" eliminado.')
    return redirect('horarios:gestion_docentes')

@admin_required
def gestion_cursos(request):
    cursos = Curso.objects.prefetch_related('paralelos').order_by('nombre')
    return render(request, 'horarios/gestion/cursos.html', {'cursos': cursos})


@admin_required
def gestion_crear_curso(request):
    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Curso creado exitosamente.')
            return redirect('horarios:gestion_cursos')
    else:
        form = CursoForm()
    return render(request, 'horarios/gestion/curso_form.html', {
        'form': form, 'titulo': 'Crear Curso', 'accion': 'Crear',
    })


@admin_required
def gestion_editar_curso(request, curso_id):
    curso = get_object_or_404(Curso, pk=curso_id)
    if request.method == 'POST':
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Curso actualizado exitosamente.')
            return redirect('horarios:gestion_cursos')
    else:
        form = CursoForm(instance=curso)
    return render(request, 'horarios/gestion/curso_form.html', {
        'form': form, 'titulo': f'Editar — {curso.nombre}', 'accion': 'Guardar cambios',
    })


@admin_required
@require_POST
def gestion_eliminar_curso(request, curso_id):
    curso = get_object_or_404(Curso, pk=curso_id)
    nombre = curso.nombre
    curso.delete()
    messages.success(request, f'Curso "{nombre}" eliminado.')
    return redirect('horarios:gestion_cursos')


@admin_required
def gestion_crear_paralelo(request):
    curso_init = request.GET.get('curso')
    if request.method == 'POST':
        form = ParaleloForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Paralelo creado exitosamente.')
            return redirect('horarios:gestion_cursos')
    else:
        form = ParaleloForm(initial={'curso': curso_init} if curso_init else {})
    return render(request, 'horarios/gestion/paralelo_form.html', {
        'form': form, 'titulo': 'Crear Paralelo', 'accion': 'Crear',
    })


@admin_required
def gestion_editar_paralelo(request, paralelo_id):
    paralelo = get_object_or_404(Paralelo, pk=paralelo_id)
    if request.method == 'POST':
        form = ParaleloForm(request.POST, instance=paralelo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Paralelo actualizado exitosamente.')
            return redirect('horarios:gestion_cursos')
    else:
        form = ParaleloForm(instance=paralelo)
    return render(request, 'horarios/gestion/paralelo_form.html', {
        'form': form, 'titulo': f'Editar — {paralelo}', 'accion': 'Guardar cambios',
    })


@admin_required
@require_POST
def gestion_eliminar_paralelo(request, paralelo_id):
    paralelo = get_object_or_404(Paralelo, pk=paralelo_id)
    nombre = str(paralelo)
    paralelo.delete()
    messages.success(request, f'Paralelo "{nombre}" eliminado.')
    return redirect('horarios:gestion_cursos')

@admin_required
def gestion_asignaturas(request):
    asignaturas = Asignatura.objects.all().order_by('nombre')
    return render(request, 'horarios/gestion/asignaturas.html', {'asignaturas': asignaturas})

@admin_required
def gestion_crear_asignatura(request):
    if request.method == 'POST':
        form = AsignaturaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignatura creada exitosamente.')
            return redirect('horarios:gestion_asignaturas')
    else:
        form = AsignaturaForm()
    return render(request, 'horarios/gestion/asignatura_form.html', {
        'form': form, 'titulo': 'Crear Asignatura', 'accion': 'Crear',
    })

@admin_required
def gestion_editar_asignatura(request, asignatura_id):
    asignatura = get_object_or_404(Asignatura, pk=asignatura_id)
    if request.method == 'POST':
        form = AsignaturaForm(request.POST, instance=asignatura)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignatura actualizada exitosamente.')
            return redirect('horarios:gestion_asignaturas')
    else:
        form = AsignaturaForm(instance=asignatura)
    return render(request, 'horarios/gestion/asignatura_form.html', {
        'form': form, 'titulo': f'Editar — {asignatura.nombre}', 'accion': 'Guardar cambios',
    })

@admin_required
@require_POST
def gestion_eliminar_asignatura(request, asignatura_id):
    asignatura = get_object_or_404(Asignatura, pk=asignatura_id)
    nombre = asignatura.nombre
    asignatura.delete()
    messages.success(request, f'Asignatura "{nombre}" eliminada.')
    return redirect('horarios:gestion_asignaturas')

# ──────────────────────────────────────────────────────────────────────────────
# MÓDULO GENERADOR AUTOMÁTICO
# ──────────────────────────────────────────────────────────────────────────────

@admin_required
def gestion_lecciones(request):
    lecciones_qs = Leccion.objects.select_related('docente__usuario', 'asignatura', 'curso', 'paralelo', 'aula_requerida').all().order_by('docente__usuario__first_name', 'asignatura__nombre', 'curso__nombre', 'paralelo__identificador')
    
    agrupadas = {}
    for l in lecciones_qs:
        key = (l.docente_id, l.asignatura_id)
        if key not in agrupadas:
            agrupadas[key] = {
                'leccion_base': l,
                'docente': l.docente,
                'asignatura': l.asignatura,
                'cursos': set(),
                'ids': [],
                'horas_totales': 0
            }
        agrupadas[key]['cursos'].add(l.curso.nombre)
        agrupadas[key]['ids'].append(l.id)
        agrupadas[key]['horas_totales'] += l.horas_semanales
        
    lecciones_list = []
    for data in agrupadas.values():
        leccion = data['leccion_base']
        docente = data['docente']
        
        # Pluralizar cursos (ej: Octavo -> Octavos)
        cursos_plural = []
        for c in sorted(data['cursos']):
            if c.endswith('o') or c.endswith('a'):
                cursos_plural.append(c + 's')
            else:
                cursos_plural.append(c)
                
        lecciones_list.append({
            'id': leccion.id,
            'docente': docente,
            'asignaturas_str': data['asignatura'].nombre,
            'horas_totales': data['horas_totales'],
            'es_grupo': len(data['ids']) > 1,
            'paralelos_str': ", ".join(cursos_plural),
            'curso': leccion.curso,
            'paralelo': leccion.paralelo,
            'ids_grupo': ",".join(map(str, data['ids']))
        })
        
    return render(request, 'horarios/gestion/lecciones.html', {'lecciones': lecciones_list})

from django.db.models import Q

@admin_required
def gestion_crear_leccion(request):
    docente_preseleccionado = request.GET.get('docente')
    if request.method == 'POST':
        docente_id = request.POST.get('docente')
        if not docente_id:
            messages.error(request, 'Debe seleccionar un docente.')
            return redirect('horarios:gestion_crear_leccion')

        num_blocks = int(request.POST.get('num_blocks', 0))
        aula_id = request.POST.get('aula_requerida')
        max_horas = request.POST.get('max_horas_seguidas', 2)
        permitir_doble = request.POST.get('permitir_doble') == 'on'
        dias_separados = request.POST.get('dias_separados') == 'on'

        creadas = 0
        razon_salto = []

        for i in range(1, num_blocks + 1):
            asignatura_id = request.POST.get(f'asignatura_{i}')
            horas = request.POST.get(f'horas_{i}')
            paralelos_ids = request.POST.getlist(f'paralelos_{i}')

            if not asignatura_id:
                razon_salto.append(f'Fila {i}: no se seleccionó asignatura.')
                continue
            if not horas:
                razon_salto.append(f'Fila {i}: no se indicaron horas.')
                continue
            if not paralelos_ids:
                razon_salto.append(f'Fila {i}: no se seleccionó ningún paralelo.')
                continue

            for paralelo_id in paralelos_ids:
                try:
                    paralelo = Paralelo.objects.get(pk=paralelo_id)
                    Leccion.objects.update_or_create(
                        docente_id=docente_id,
                        asignatura_id=asignatura_id,
                        paralelo_id=paralelo_id,
                        defaults={
                            'curso_id': paralelo.curso_id,
                            'aula_requerida_id': aula_id if aula_id else None,
                            'horas_semanales': horas,
                            'max_horas_seguidas': max_horas,
                            'permitir_doble': permitir_doble,
                            'dias_separados': dias_separados
                        }
                    )
                    creadas += 1
                except Paralelo.DoesNotExist:
                    pass

        if creadas > 0:
            messages.success(request, f'Se crearon/actualizaron {creadas} lecciones exitosamente.')
            if 'guardar_y_otro' in request.POST:
                url = reverse('horarios:gestion_crear_leccion')
                return redirect(url)
            return redirect('horarios:gestion_lecciones')
        else:
            detalles = ' | '.join(razon_salto) if razon_salto else 'No se enviaron datos.'
            messages.warning(request, f'No se procesó ninguna lección. {detalles}')
            url = reverse('horarios:gestion_crear_leccion') + '?docente=' + str(docente_id)
            return redirect(url)
            
    else:
        form = LeccionForm()
    return render(request, 'horarios/gestion/leccion_form.html', {
        'form': form, 'titulo': 'Crear Lección', 'accion': 'Crear',
        'docente_preseleccionado': docente_preseleccionado,
        'paralelos_json': _paralelos_json(),
        'todos_paralelos_json': _todos_paralelos_json(),
        'asignaturas_all': Asignatura.objects.order_by('nombre'),
        'aulas_all': Aula.objects.all(),
        'cursos_all': Curso.objects.all(),
        'paralelos_all': Paralelo.objects.all(),
    })

@admin_required
def gestion_editar_leccion(request, leccion_id):
    leccion = get_object_or_404(Leccion, pk=leccion_id)
    if request.method == 'POST':
        form = LeccionForm(request.POST, instance=leccion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Lección actualizada.')
            return redirect('horarios:gestion_lecciones')
    else:
        form = LeccionForm(instance=leccion)
    return render(request, 'horarios/gestion/leccion_form.html', {
        'form': form, 'titulo': 'Editar Lección', 'accion': 'Guardar',
        'paralelos_json': _paralelos_json()
    })

@admin_required
@require_POST
def gestion_eliminar_leccion(request, leccion_id):
    leccion = get_object_or_404(Leccion, pk=leccion_id)
    leccion.delete()
    messages.success(request, 'Lección eliminada.')
    return redirect('horarios:gestion_lecciones')

@admin_required
@require_POST
def gestion_eliminar_lecciones_masivo(request):
    ids_str = request.POST.get('lecciones_ids', '')
    if ids_str:
        ids = [int(i) for i in ids_str.split(',') if i.strip().isdigit()]
        count = Leccion.objects.filter(id__in=ids).count()
        if count > 0:
            Leccion.objects.filter(id__in=ids).delete()
            messages.success(request, f'Se eliminaron {count} lecciones.')
        else:
            messages.warning(request, 'No se encontraron las lecciones especificadas.')
    else:
        messages.warning(request, 'No se seleccionó ninguna lección.')
    return redirect('horarios:gestion_lecciones')

@admin_required
@require_POST
def gestion_eliminar_todas_lecciones(request):
    count = Leccion.objects.count()
    if count > 0:
        Leccion.objects.all().delete()
        messages.success(request, f'Se eliminaron todas las lecciones ({count} en total).')
    else:
        messages.warning(request, 'No hay lecciones para eliminar.')
    return redirect('horarios:gestion_lecciones')


@admin_required
def gestion_disponibilidad(request):
    docente_id = request.GET.get('docente')
    docentes = Docente.objects.select_related('usuario').all().order_by('usuario__first_name')
    
    context = {'docentes': docentes, 'docente_seleccionado': None, 'dias': Horario.DIAS, 'periodos': PERIODOS_ROWS}
    
    if docente_id:
        docente = get_object_or_404(Docente, pk=docente_id)
        bloqueos = DisponibilidadDocente.objects.filter(docente=docente)
        
        # Mapa: (dia, hora_inicio_str) -> tipo
        mapa_bloqueos = {}
        for b in bloqueos:
            mapa_bloqueos[(b.dia.lower(), b.hora_inicio.strftime('%H:%M:%S'))] = b.tipo
            
        context['docente_seleccionado'] = docente
        context['mapa_bloqueos'] = mapa_bloqueos
        
    return render(request, 'horarios/gestion/disponibilidad.html', context)


@admin_required
@require_POST
def api_guardar_disponibilidad(request):
    import json
    data = json.loads(request.body)
    docente_id = data.get('docente_id')
    dia = data.get('dia')
    hora_inicio = data.get('hora_inicio')
    hora_fin = data.get('hora_fin')
    tipo = data.get('tipo')  # 'bloqueado', 'preferido' o 'libre'
    
    docente = get_object_or_404(Docente, pk=docente_id)
    
    if tipo == 'libre':
        DisponibilidadDocente.objects.filter(docente=docente, dia=dia, hora_inicio=hora_inicio).delete()
    else:
        obj, created = DisponibilidadDocente.objects.update_or_create(
            docente=docente, dia=dia, hora_inicio=hora_inicio,
            defaults={'hora_fin': hora_fin, 'tipo': tipo}
        )
        
    return JsonResponse({'status': 'ok'})


@admin_required
def gestion_generador(request):
    lecciones_count = Leccion.objects.count()
    horas_totales = sum(l.horas_semanales for l in Leccion.objects.all())
    docentes_count = Leccion.objects.values('docente').distinct().count()
    
    sesiones_recientes = SesionGenerador.objects.order_by('-creado_en')[:5]
    
    context = {
        'lecciones_count': lecciones_count,
        'horas_totales': horas_totales,
        'docentes_count': docentes_count,
        'sesiones_recientes': sesiones_recientes
    }
    return render(request, 'horarios/gestion/generador.html', context)

@admin_required
@require_POST
def api_verificar_coherencia(request):
    from .scheduler import _cargar_datos_generacion, _verificar_coherencia
    import asyncio
    
    # Para ser simple y síncrono en la vista:
    lecciones = list(Leccion.objects.all())
    disp = DisponibilidadDocente.objects.filter(tipo='bloqueado')
    bloqueos = []
    for d in disp:
        dia_idx = DAY_TO_COL.get(d.dia.lower(), 2) - 2 # Mapeo rápido a 0-4
        hora_str = d.hora_inicio.strftime("%H:%M:%S")
        per_idx = next((i for i, r in enumerate(PERIODOS_ROWS) if r[1] == hora_str), None)
        if per_idx is not None:
            bloqueos.append({'docente_id': d.docente_id, 'dia_idx': dia_idx, 'per_idx': per_idx})
            
    valido, advertencias = _verificar_coherencia(lecciones, bloqueos)
    return JsonResponse({'valido': valido, 'advertencias': advertencias})

@admin_required
@require_POST
def api_ejecutar_generador(request):
    import json
    data = json.loads(request.body)
    modo = data.get('modo', 'estandar')
    max_intentos = 2000000 if modo == 'complejo' else 50000
    
    sesion = SesionGenerador.objects.create(modo=modo)
    
    # Lanzar tarea asíncrona en segundo plano real requiere Celery o threading.
    # Dado que estamos usando python estándar sin dependencias fuertes de workers,
    # usaremos un Thread simple por practicidad.
    import threading
    import asyncio
    from .scheduler import generar_horario_async
    
    def run_async_task(sid, max_int):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(generar_horario_async(sid, max_int))
        loop.close()
        
    t = threading.Thread(target=run_async_task, args=(str(sesion.sesion_id), max_intentos))
    t.daemon = True
    t.start()
    
    return JsonResponse({'status': 'ok', 'sesion_id': str(sesion.sesion_id)})

@admin_required
def api_estado_generador(request, sesion_id):
    sesion = get_object_or_404(SesionGenerador, sesion_id=sesion_id)
    return JsonResponse({
        'estado': sesion.estado,
        'advertencias': sesion.advertencias,
        'sin_asignar': sesion.sin_asignar,
        'publicado': sesion.publicado
    })

@admin_required
def gestion_borrador(request, sesion_id):
    sesion = get_object_or_404(SesionGenerador, sesion_id=sesion_id)
    borradores = sesion.borradores.select_related(
        'leccion__docente__usuario', 'leccion__asignatura',
        'leccion__curso', 'leccion__paralelo', 'aula'
    ).order_by('dia', 'hora_inicio')
    
    # Adaptamos borradores al formato de grid
    base_items = []
    for b in borradores:
        start_str = b.hora_inicio.strftime("%H:%M:%S")
        end_str = b.hora_fin.strftime("%H:%M:%S")
        
        row_start = TIME_TO_ROW.get(start_str, 2)
        row_end = TIME_TO_ROW.get(end_str, row_start + 1)
        col = DAY_TO_COL.get(b.dia.lower(), 2)

        clave_color = f"{b.leccion.asignatura.nombre}|{b.leccion.curso_id}|{b.leccion.paralelo_id}"
        colors = get_color_for_materia(clave_color)
        
        # Mapear atributos para reusar schedule_item.html
        b.asignatura = b.leccion.asignatura
        b.docente = b.leccion.docente
        b.curso = b.leccion.curso
        b.paralelo = b.leccion.paralelo
        b.id_borrador = b.id
        
        base_items.append({
            'horario': b,
            'grid_row_start': row_start,
            'grid_row_end': row_end,
            'grid_col': col,
            'color_bg':     colors['bg'],
            'color_border': colors['border'],
            'color_text':   colors['text'],
        })

    # Fusionar bloques contiguos (doble período visual)
    base_items.sort(key=lambda x: (x['grid_col'], x['grid_row_start']))
    merged = []
    for item in base_items:
        if not merged:
            merged.append(item)
            continue
        last = merged[-1]
        can_merge = (
            last['grid_col'] == item['grid_col'] and
            last['grid_row_end'] == item['grid_row_start'] and
            last['horario'].asignatura.id == item['horario'].asignatura.id and
            last['horario'].docente.id == item['horario'].docente.id and
            last['horario'].curso.id == item['horario'].curso.id and
            last['horario'].paralelo.id == item['horario'].paralelo.id
        )
        if can_merge:
            last['grid_row_end'] = item['grid_row_end']
            # Actualizar hora_fin para el badge de horario
            last['horario'].hora_fin = item['horario'].hora_fin
        else:
            merged.append(item)
        
    context = {
        'sesion': sesion,
        'items': merged,
        'dias': Horario.DIAS,
        'periodos': PERIODOS_ROWS
    }
    return render(request, 'horarios/gestion/borrador.html', context)

@admin_required
@require_POST
def api_publicar_borrador(request, sesion_id):
    from .scheduler import publicar_horario
    nuevos, grupos = publicar_horario(sesion_id)
    return JsonResponse({'status': 'ok', 'nuevos': nuevos, 'grupos': grupos})

@admin_required
@require_POST
def api_descartar_borrador(request, sesion_id):
    try:
        sesion = SesionGenerador.objects.get(sesion_id=sesion_id)
        sesion.delete()
        return JsonResponse({'status': 'ok'})
    except SesionGenerador.DoesNotExist:
        return JsonResponse({'status': 'ok', 'msg': 'La sesión ya no existe.'})  
    except Exception as e:
        return JsonResponse({'status': 'error', 'msg': str(e)}, status=500)


@admin_required
def gestion_atencion_padres(request):
    horarios_atencion = Horario.objects.filter(tipo='atencion').select_related(
        'docente__usuario'
    ).order_by('dia', 'hora_inicio')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            form = AtencionPadresForm(request.POST)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.tipo = 'atencion'
                obj.save()
                messages.success(request, "Horario de atención añadido correctamente.")
                return redirect('horarios:gestion_atencion_padres')
            else:
                messages.error(request, "Error al crear el horario. Revisa los campos.")
        elif action == 'edit':
            obj_id = request.POST.get('horario_id')
            obj = get_object_or_404(Horario, id=obj_id, tipo='atencion')
            form = AtencionPadresForm(request.POST, instance=obj)
            if form.is_valid():
                form.save()
                messages.success(request, "Horario de atención modificado correctamente.")
                return redirect('horarios:gestion_atencion_padres')
            else:
                messages.error(request, "Error al modificar el horario.")
        elif action == 'delete':
            obj_id = request.POST.get('horario_id')
            obj = get_object_or_404(Horario, id=obj_id, tipo='atencion')
            obj.delete()
            messages.success(request, "Horario de atención eliminado.")
            return redirect('horarios:gestion_atencion_padres')
        elif action == 'move_all':
            dia_origen = request.POST.get('dia_origen', '').strip().lower()
            dia_destino = request.POST.get('dia_destino', '').strip().lower()
            dias_validos = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
            if dia_origen in dias_validos and dia_destino in dias_validos and dia_origen != dia_destino:
                count = Horario.objects.filter(tipo='atencion', dia=dia_origen).update(dia=dia_destino)
                messages.success(request, f"Se movieron {count} horario(s) de {dia_origen.capitalize()} a {dia_destino.capitalize()}.")
            else:
                messages.error(request, "Selecciona un día de destino válido diferente al de origen.")
            return redirect('horarios:gestion_atencion_padres')

    form = AtencionPadresForm()
    
    dias_list = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
    horarios_por_dia = {dia: [] for dia in dias_list}
    for h in horarios_atencion:
        if h.dia in horarios_por_dia:
            horarios_por_dia[h.dia].append(h)
            
    # Convertir a lista de tuplas para mantener el orden de días
    dias_ordenados = [(dia, dia.capitalize(), horarios_por_dia[dia]) for dia in dias_list]
    
    context = {
        'titulo_seccion': 'Atención a Padres de Familia',
        'horarios': horarios_atencion,
        'dias_ordenados': dias_ordenados,
        'form': form,
        'dias_choices': Horario.DIAS,
    }
    return render(request, 'horarios/gestion/atencion_padres.html', context)


@login_required
def atencion_padres_publico(request):
    horarios_qs = Horario.objects.filter(tipo='atencion').select_related(
        'docente__usuario'
    ).order_by('dia', 'hora_inicio', 'docente__usuario__first_name')
    
    dias_list = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
    horarios_por_dia = {dia: [] for dia in dias_list}
    for h in horarios_qs:
        if h.dia in horarios_por_dia:
            horarios_por_dia[h.dia].append(h)
    
    # Lista de tuplas (dia_id, dia_nombre, horarios) — solo los que tienen datos
    dias_con_datos = [(dia, dia.capitalize(), horarios_por_dia[dia]) for dia in dias_list if horarios_por_dia[dia]]
    # Todos los días para la cuadrícula
    dias_ordenados = [(dia, dia.capitalize(), horarios_por_dia[dia]) for dia in dias_list]
    
    dias_disponibles = [d[1] for d in dias_con_datos]
    if len(dias_disponibles) == 0:
        dias_texto = "No hay horarios disponibles"
    elif len(dias_disponibles) == 1:
        dias_texto = dias_disponibles[0]
    elif len(dias_disponibles) == 2:
        dias_texto = f"{dias_disponibles[0]} y {dias_disponibles[1]}"
    elif len(dias_disponibles) == 5:
        dias_texto = "Lunes a Viernes"
    else:
        dias_texto = ", ".join(dias_disponibles[:-1]) + " y " + dias_disponibles[-1]
    
    # Detectar si hoy hay atención
    hoy = datetime.datetime.now().weekday()
    dias_semana_map = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
    dia_hoy = dias_semana_map[hoy]
    hay_atencion_hoy = dia_hoy in horarios_por_dia and len(horarios_por_dia.get(dia_hoy, [])) > 0
        
    return render(request, 'horarios/atencion_padres_publico.html', {
        'dias_ordenados': dias_ordenados,
        'dias_texto': dias_texto,
        'hay_atencion_hoy': hay_atencion_hoy,
        'dia_hoy': dia_hoy,
        'total_docentes': horarios_qs.values('docente').distinct().count(),
    })

def send_mass_html_email_thread(subject, text_content, template_name, context, recipient_list):
    html_content = render_to_string(template_name, context)
    connection = get_connection()
    messages = []
    
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_dolorosa.png')
    logo_data = None
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_data = f.read()

    from_email_formatted = f"GeoCampus La Dolorosa <{settings.EMAIL_HOST_USER}>"

    # Enviar correos de forma individual pero sobre la misma conexión SMTP (evita Spam por BCC)
    for correo in recipient_list:
        if not correo:
            continue
            
        msg = EmailMultiAlternatives(
            subject=subject,
            body=text_content,
            from_email=from_email_formatted,
            to=[correo],
            reply_to=[settings.EMAIL_HOST_USER]
        )
        msg.attach_alternative(html_content, "text/html")
        
        if logo_data:
            logo_img = MIMEImage(logo_data)
            logo_img.add_header('Content-ID', '<logo>')
            logo_img.add_header('Content-Disposition', 'inline', filename='logo_dolorosa.png')
            msg.attach(logo_img)
            
        messages.append(msg)
        
    try:
        connection.send_messages(messages)
    except Exception as e:
        print(f"Error sending mass email: {e}")

@login_required
@admin_required
def enviar_notificacion_padres_view(request):
    dias_semana_map = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
    dia_hoy = dias_semana_map[datetime.datetime.now().weekday()].capitalize()
    
    correos = list(Usuario.objects.filter(rol='estudiante').exclude(email='').values_list('email', flat=True))
    site_url = "https://geocampus.vercel.app" + reverse('horarios:atencion_padres_publico')
    
    send_mass_html_email_thread(
        f"Hoy hay Atención a Padres de Familia - {dia_hoy}",
        f"El día de hoy ({dia_hoy}) tenemos jornada de Atención a Padres de Familia en la Unidad Educativa La Dolorosa. Revise los horarios en: {site_url}",
        'horarios/emails/notificacion_padres.html',
        {'dia_hoy': dia_hoy, 'site_url': site_url},
        correos
    )
    
    messages.success(request, f'Se ha iniciado el envío del boletín a {len(correos)} usuarios registrados. Esto puede tardar un par de minutos en completarse.')
    return redirect('horarios:gestion_atencion_padres')

def cron_notificar_padres_view(request):
    key = request.GET.get('key')
    if key != getattr(settings, 'CRON_SECRET_KEY', 'default-cron-secret-2026'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)
        
    dias_semana_map = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
    dia_hoy_str = dias_semana_map[datetime.datetime.now().weekday()]
    
    hay_atencion_hoy = Horario.objects.filter(tipo='atencion', dia=dia_hoy_str).exists()
    
    if not hay_atencion_hoy:
        return JsonResponse({'status': 'ok', 'message': f'No hay atencion a padres el {dia_hoy_str}.'})
        
    dia_hoy_capitalized = dia_hoy_str.capitalize()
    correos = list(Usuario.objects.filter(rol='estudiante').exclude(email='').values_list('email', flat=True))
    site_url = "https://geocampus.vercel.app" + reverse('horarios:atencion_padres_publico')
    
    send_mass_html_email_thread(
        f"Hoy hay Atención a Padres de Familia - {dia_hoy_capitalized}",
        f"El día de hoy ({dia_hoy_capitalized}) tenemos jornada de Atención a Padres de Familia en la Unidad Educativa La Dolorosa. Revise los horarios en: {site_url}",
        'horarios/emails/notificacion_padres.html',
        {'dia_hoy': dia_hoy_capitalized, 'site_url': site_url},
        correos
    )
    
    return JsonResponse({'status': 'ok', 'message': f'Notificaciones enviadas a {len(correos)} usuarios.'})
